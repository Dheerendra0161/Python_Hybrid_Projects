import os
import random

import openpyxl


def read_excel_data(sheet_name):
    # file_path = 'utilities/excel_data_read.xlsx'
    current_dir = os.path.dirname(os.path.realpath(__file__))

    # Construct the file path
    file_path = os.path.join(current_dir, 'excel_data_read.xlsx')
    # Load the workbook and select the sheet
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]

    max_row = sheet.max_row

    # Read data starting from the second row (skipping headers)
    data = []
    for row in range(2, max_row + 1):
        cell_value_1 = sheet.cell(row=row, column=1).value
        data.append(cell_value_1)

    workbook.close()
    return data


# # Call the function and store the returned data
# returned_data = read_excel_data("Relations")
#
# # Print the returned data
# for row_data in returned_data:
#     print(row_data)


def read_excel_data1(sheet_name, first_name_col, last_name_col, email_col, pass_col, mobile_col):
    # Get the directory of the current script
    current_dir = os.path.dirname(os.path.realpath(__file__))

    # Construct the file path
    file_path = os.path.join(current_dir, 'excel_data_read.xlsx')

    # Load the workbook and select the sheet
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]

    max_row = sheet.max_row

    # Read data starting from the second row (skipping headers)
    data = []
    for row in range(2, max_row + 1):
        first_name = sheet.cell(row=row, column=first_name_col).value
        last_name = sheet.cell(row=row, column=last_name_col).value
        email = sheet.cell(row=row, column=email_col).value
        password = sheet.cell(row=row, column=pass_col).value
        mobile = sheet.cell(row=row, column=mobile_col).value

        # Create a dictionary for each row
        row_data = {
            "firstName": first_name,
            "lastName": last_name,
            "email": email,
            # "password": password,
            # "mobile": mobile
        }

        # Append the row data to the list
        data.append(row_data)

    workbook.close()

    return data


# data = read_excel_data1("Relations", 1, 2, 3, 4, 5)
# print(data)


def read_excel_data2(sheet_name, *column_names):
    current_dir = os.path.dirname(os.path.realpath(__file__))
    file_path = os.path.join(current_dir, 'excel_data_read.xlsx')
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    max_row = sheet.max_row
    column_indices = {col_name: idx + 1 for idx, col_name in enumerate(column_names)}
    data = []
    for row in range(2, max_row + 1):
        # Create a dictionary for each row
        row_data = {}
        for col_name in column_indices:
            cell_value = sheet.cell(row=row, column=column_indices[col_name]).value
            row_data[col_name] = cell_value

        # Append the row data to the list
        data.append(row_data)

    workbook.close()

    return data


# data = read_excel_data2("Relations6", 1)
# print(data)


# Read the data from the excel sheet by giving the desired column value


def read_excel_data3(sheet_name, column_number):
    # Get the directory of the current script
    current_dir = os.path.dirname(os.path.realpath(__file__))
    file_path = os.path.join(current_dir, 'excel_data_read.xlsx')
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    max_columns = sheet.max_column
    returned_data = []
    for row in range(2, sheet.max_row + 1):
        cell_value = sheet.cell(row=row, column=column_number).value
        if cell_value is not None:
            returned_data.append(cell_value)

    workbook.close()

    return returned_data, max_columns


# returned_data, max_columns = read_excel_data3("Insurer_Relation", 7)  # Example: Reading data from column 2 (B)
# print("Data:", returned_data)
# print("Total Columns:", max_columns)


def read_excel_data5(sheet_name, target_heading):
    # Get the directory of the current script
    current_dir = os.path.dirname(os.path.realpath(__file__))

    # Construct the file path
    file_path = os.path.join(current_dir, 'excel_data_read.xlsx')

    # Load the workbook and select the sheet
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]

    # Find the column number of the target_heading
    target_col = None
    for col in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(row=1, column=col).value
        if cell_value == target_heading:
            target_col = col
            break

    if target_col is None:
        raise ValueError(f"Column with heading '{target_heading}' not found in the sheet")

    # Read all cell values under the target_heading
    data = []
    for row in range(2, sheet.max_row + 1):
        cell_value = sheet.cell(row=row, column=target_col).value
        if cell_value is not None:
            data.append(cell_value)

    workbook.close()

    return data


# Example usage:
# sheet_name = "Insurer_Relation"
# # target_heading = "Relation3"
# random_data_relation = ['Relation', 'Relation1', "Relation2", "Relation3", "Relation4", "Relation5"]
# random_data =random.choice(random_data_relation)
# data = read_excel_data5(sheet_name, random_data)
# print(data)
# returned_data = read_excel_data5("Cities", "parents_cities")
# parents_city_name = random.choice(returned_data)
# print(parents_city_name)